"""Deterministic verifier: confirm suspected errors by independent recompute.

Two check families, each ending in a numeric recompute-and-compare so every
finding is backed by 'actual value vs. value we independently derived':

  Check A (row-pattern): E3/E4/E5 -- a cell in a horizontal band whose formula
    breaks the band's consensus template. Recompute under the consensus formula;
    report only if the shown value actually differs.

  Check B (totals): E1/E7 -- a total cell that should equal SUM of its row band.
    Recompute the sum; report if the shown value differs.

This module is pure code (no LLM). In the full system the Hunter agent proposes
candidates and this verifies them; for the spike the checks also do the finding,
which is enough to prove the recompute kernel catches seeded errors.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Optional

from openpyxl.utils import get_column_letter, column_index_from_string

from workbook import Workbook
from formula_engine import _CELL_RE

TOL = 1e-6


@dataclass
class Finding:
    cell: str
    error_class: str
    actual: float
    expected: float
    detail: str


_ANCHORED = re.compile(r"(\$?)([A-Z]+)(\$?)(\d+)")


def to_template(formula: str, coord: str) -> str:
    """Canonicalize a formula relative to `coord` using offset tokens, so two
    cells with the 'same' copied formula produce identical template strings.
    Relative col -> {c:+d}, relative row -> {r:+d}; absolute parts kept literal.
    """
    cm = re.fullmatch(r"([A-Z]+)(\d+)", coord)
    base_c = column_index_from_string(cm.group(1))
    base_r = int(cm.group(2))

    def repl(m: re.Match) -> str:
        dollar_c, col, dollar_r, row = m.groups()
        c_tok = col if dollar_c else f"{{c:{column_index_from_string(col) - base_c:+d}}}"
        r_tok = row if dollar_r else f"{{r:{int(row) - base_r:+d}}}"
        return f"{dollar_c}{c_tok}{dollar_r}{r_tok}"

    return _ANCHORED.sub(repl, formula)


def from_template(template: str, coord: str) -> str:
    """Instantiate an offset template into a concrete formula at `coord`."""
    cm = re.fullmatch(r"([A-Z]+)(\d+)", coord)
    base_c = column_index_from_string(cm.group(1))
    base_r = int(cm.group(2))

    def c_repl(m: re.Match) -> str:
        return get_column_letter(base_c + int(m.group(1)))

    def r_repl(m: re.Match) -> str:
        return str(base_r + int(m.group(1)))

    out = re.sub(r"\{c:([+-]\d+)\}", c_repl, template)
    out = re.sub(r"\{r:([+-]\d+)\}", r_repl, out)
    return out


def _coord(col_i: int, row_i: int) -> str:
    return f"{get_column_letter(col_i)}{row_i}"


def check_row_pattern(wb: Workbook, values: dict, row: int,
                      col_start: int, col_end: int) -> list[Finding]:
    """Check A: find cells in a row band that break the consensus formula template."""
    from formula_engine import evaluate
    templates: dict[str, list[int]] = {}
    for c in range(col_start, col_end + 1):
        cell = wb.cells.get(_coord(c, row))
        if cell and cell.formula:
            tpl = to_template(cell.formula, _coord(c, row))
            templates.setdefault(tpl, []).append(c)
    if len(templates) < 2:
        return []  # all consistent (or nothing to compare)

    consensus = max(templates.items(), key=lambda kv: len(kv[1]))[0]
    findings = []
    for tpl, cols in templates.items():
        if tpl == consensus:
            continue
        for c in cols:
            coord = _coord(c, row)
            expected_formula = from_template(consensus, coord)
            try:
                expected = evaluate(expected_formula, values)
            except Exception:
                # Expected formula references a non-existent/label cell -> this
                # cell is a legitimate band edge (e.g. first col of a running
                # total), not a verifiable error. Skip.
                continue
            actual = values[coord]
            if abs(actual - expected) > TOL:
                findings.append(Finding(
                    coord, "E3", actual, expected,
                    f"formula breaks row {row} pattern; "
                    f"expected like '{expected_formula}'",
                ))
    return findings


def check_total(wb: Workbook, values: dict, total_cell: str,
                row: int, col_start: int, col_end: int) -> Optional[Finding]:
    """Check B: a total cell should equal SUM of its row band (horizontal)."""
    from formula_engine import evaluate
    band = f"=SUM({_coord(col_start, row)}:{_coord(col_end, row)})"
    return _compare_total(wb, values, total_cell, band)


def check_col_total(wb: Workbook, values: dict, total_cell: str,
                    col_i: int, row_start: int, row_end: int) -> Optional[Finding]:
    """Check B (vertical): a grand-total cell should equal SUM of its column band."""
    band = f"=SUM({_coord(col_i, row_start)}:{_coord(col_i, row_end)})"
    return _compare_total(wb, values, total_cell, band)


def _compare_total(wb: Workbook, values: dict, total_cell: str, band: str) -> Optional[Finding]:
    from formula_engine import evaluate
    expected = evaluate(band, values)
    actual = values[total_cell]
    if abs(actual - expected) > TOL:
        cell = wb.cells.get(total_cell)
        cls = "E7" if (cell and cell.formula is None) else "E1"
        return Finding(total_cell, cls, actual, expected,
                       f"total != SUM of band; expected {band}")
    return None
