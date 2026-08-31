"""Generate one seeded financial model + ground-truth file.

Builds a 12-month SaaS P&L projection with clean, consistent formulas, then
injects a known set of errors (E1, E3, E7). Writes:
  data/sheets/model_01.xlsx   (the errored workbook the auditor will inspect)
  data/truth/model_01.json    (ground truth: injected cells + class + fix)
"""
from __future__ import annotations

import json
import os

import openpyxl
from openpyxl.utils import get_column_letter

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# columns B..M = 12 months, N = total
FIRST_COL = 2          # B
LAST_COL = FIRST_COL + 11  # M
TOTAL_COL = LAST_COL + 1   # N

# Rows
R_NEW = 2       # New customers (input)
R_CUM = 3       # Cumulative customers = prev + new
R_PRICE = 4     # Price per customer (input, constant)
R_REV = 5       # Revenue = cumulative * price
R_UCOST = 6     # Unit cost (input, constant)
R_COST = 7      # Total cost = cumulative * unit cost
R_PROFIT = 8    # Gross profit = revenue - cost


def col(i: int) -> str:
    return get_column_letter(i)


def build_clean(ws) -> None:
    ws["A1"] = "SaaS 12-Month P&L Projection"
    ws["A2"] = "New Customers"
    ws["A3"] = "Cumulative Customers"
    ws["A4"] = "Price / Customer"
    ws["A5"] = "Revenue"
    ws["A6"] = "Unit Cost / Customer"
    ws["A7"] = "Total Cost"
    ws["A8"] = "Gross Profit"
    ws[f"{col(TOTAL_COL)}1"] = "FY Total"

    new_customers = [50, 60, 72, 80, 95, 110, 130, 150, 175, 200, 230, 260]
    for idx, c in enumerate(range(FIRST_COL, LAST_COL + 1)):
        letter = col(c)
        ws[f"{letter}1"] = MONTHS[idx]
        ws[f"{letter}{R_NEW}"] = new_customers[idx]
        # cumulative
        if c == FIRST_COL:
            ws[f"{letter}{R_CUM}"] = f"={letter}{R_NEW}"
        else:
            prev = col(c - 1)
            ws[f"{letter}{R_CUM}"] = f"={prev}{R_CUM}+{letter}{R_NEW}"
        ws[f"{letter}{R_PRICE}"] = 40
        ws[f"{letter}{R_REV}"] = f"={letter}{R_CUM}*{letter}{R_PRICE}"
        ws[f"{letter}{R_UCOST}"] = 15
        ws[f"{letter}{R_COST}"] = f"={letter}{R_CUM}*{letter}{R_UCOST}"
        ws[f"{letter}{R_PROFIT}"] = f"={letter}{R_REV}-{letter}{R_COST}"

    # FY totals
    b, m = col(FIRST_COL), col(LAST_COL)
    t = col(TOTAL_COL)
    ws[f"{t}{R_REV}"] = f"=SUM({b}{R_REV}:{m}{R_REV})"
    ws[f"{t}{R_COST}"] = f"=SUM({b}{R_COST}:{m}{R_COST})"
    ws[f"{t}{R_PROFIT}"] = f"=SUM({b}{R_PROFIT}:{m}{R_PROFIT})"


def inject_errors(ws) -> list[dict]:
    """Overwrite specific cells with wrong formulas; return ground truth."""
    truth = []

    # E1: broken range -- FY Total Revenue misses December (M)
    t = col(TOTAL_COL)
    b = col(FIRST_COL)
    l = col(LAST_COL - 1)  # L, one short of M
    cell = f"{t}{R_REV}"
    ws[cell] = f"=SUM({b}{R_REV}:{l}{R_REV})"   # should end at M
    truth.append({
        "cell": cell, "class": "E1",
        "desc": "FY Total Revenue SUM range stops at L (Nov), omits M (Dec).",
        "wrong_formula": f"=SUM({b}{R_REV}:{l}{R_REV})",
        "correct_formula": f"=SUM({b}{R_REV}:{col(LAST_COL)}{R_REV})",
    })

    # E3: inconsistent copied formula -- August profit uses wrong multiplier logic
    aug = col(FIRST_COL + 7)  # I (Aug)
    cell = f"{aug}{R_PROFIT}"
    ws[cell] = f"={aug}{R_REV}-{aug}{R_COST}*0.9"  # neighbors are REV-COST
    truth.append({
        "cell": cell, "class": "E3",
        "desc": "Aug Gross Profit deviates from the row pattern (spurious *0.9 on cost).",
        "wrong_formula": f"={aug}{R_REV}-{aug}{R_COST}*0.9",
        "correct_formula": f"={aug}{R_REV}-{aug}{R_COST}",
    })

    # E7: aggregation mismatch -- FY Total Cost hardcoded to a stale number
    cell = f"{t}{R_COST}"
    ws[cell] = 24000  # literal, not a SUM; real sum differs
    truth.append({
        "cell": cell, "class": "E7",
        "desc": "FY Total Cost is a hardcoded literal, not SUM of monthly costs.",
        "wrong_formula": "24000 (literal)",
        "correct_formula": f"=SUM({b}{R_COST}:{col(LAST_COL)}{R_COST})",
    })

    return truth


def main() -> None:
    os.makedirs("data/sheets", exist_ok=True)
    os.makedirs("data/truth", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PnL"
    build_clean(ws)
    truth = inject_errors(ws)

    xlsx_path = "data/sheets/model_01.xlsx"
    truth_path = "data/truth/model_01.json"
    wb.save(xlsx_path)
    with open(truth_path, "w") as f:
        json.dump({"sheet": "PnL", "errors": truth}, f, indent=2)

    print(f"wrote {xlsx_path} with {len(truth)} seeded errors")
    print(f"wrote {truth_path}")


if __name__ == "__main__":
    main()
