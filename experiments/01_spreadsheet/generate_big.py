"""Large sheet to test whether baseline coverage degrades at scale.
40 line items x 24 months + totals (~1000 cells), 6 scattered subtle errors."""
import json, os, random
import openpyxl
from openpyxl.utils import get_column_letter as gcl

N_ITEMS, N_MONTHS = 40, 24
FIRST_COL = 2
LAST_COL = FIRST_COL + N_MONTHS - 1
TOTAL_COL = LAST_COL + 1
ROW_START = 3
ROW_TOTAL = ROW_START + N_ITEMS

def main():
    random.seed(7)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Big"
    ws["A1"] = "Large OpEx model"
    ws[f"{gcl(TOTAL_COL)}1"] = "FY Total"
    for j,c in enumerate(range(FIRST_COL, LAST_COL+1)):
        ws[f"{gcl(c)}1"] = f"M{j+1}"
    g = 1.02
    for i in range(N_ITEMS):
        r = ROW_START+i
        ws[f"A{r}"] = f"Item {i+1}"
        base = random.randint(500,4000)
        for c in range(FIRST_COL, LAST_COL+1):
            L=gcl(c)
            ws[f"{L}{r}"] = base if c==FIRST_COL else f"={gcl(c-1)}{r}*{g}"
        ws[f"{gcl(TOTAL_COL)}{r}"] = f"=SUM({gcl(FIRST_COL)}{r}:{gcl(LAST_COL)}{r})"
    ws[f"A{ROW_TOTAL}"]="TOTAL"
    for c in range(FIRST_COL, TOTAL_COL+1):
        L=gcl(c)
        ws[f"{L}{ROW_TOTAL}"]=f"=SUM({L}{ROW_START}:{L}{ROW_TOTAL-1})"
    truth=[]
    # 6 scattered subtle errors
    def add(cell,cls,desc): truth.append({"cell":cell,"class":cls,"desc":desc})
    r=ROW_START+13; ws[f"{gcl(10)}{r}"]=f"={gcl(9)}{r}*1.20"; add(f"{gcl(10)}{r}","E3","item14 M9 *1.20")
    r=ROW_START+27; ws[f"{gcl(TOTAL_COL)}{r}"]=f"=SUM({gcl(FIRST_COL)}{r}:{gcl(LAST_COL-1)}{r})"; add(f"{gcl(TOTAL_COL)}{r}","E1","item28 FY total short one month")
    r=ROW_START+5; ws[f"{gcl(16)}{r}"]=f"={gcl(15)}{r}*1.02+50"; add(f"{gcl(16)}{r}","E3","item6 M15 +50")
    c=gcl(12); ws[f"{c}{ROW_TOTAL}"]=f"=SUM({c}{ROW_START}:{c}{ROW_TOTAL-3})"; add(f"{c}{ROW_TOTAL}","E1","M11 grand total short 2 rows")
    ws[f"{gcl(TOTAL_COL)}{ROW_TOTAL}"]=1250000; add(f"{gcl(TOTAL_COL)}{ROW_TOTAL}","E7","grand FY total hardcoded")
    r=ROW_START+33; ws[f"{gcl(20)}{r}"]=f"={gcl(19)}{r}"; add(f"{gcl(20)}{r}","E3","item34 M19 missing *g")
    os.makedirs("data/sheets",exist_ok=True); os.makedirs("data/truth",exist_ok=True)
    wb.save("data/sheets/model_03_big.xlsx")
    json.dump({"sheet":"Big","errors":truth}, open("data/truth/model_03_big.json","w"), indent=2)
    ncells = (N_ITEMS+1)*(N_MONTHS+1)
    print(f"wrote model_03_big.xlsx (~{ncells} cells) with {len(truth)} errors")
    for e in truth: print(" ", e["cell"], e["class"], e["desc"])
main()
