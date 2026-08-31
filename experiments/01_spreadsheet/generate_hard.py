"""Generate a HARD seeded model: larger, with subtle errors that produce
plausible values -- the regime where eyeball/mental-math auditing fails but
exact recompute still catches everything.

Design choices that stress a single-prompt LLM:
  - scale: 15 expense line items x 12 months + subtotal col + total row
  - subtle errors that yield near-correct values:
      * mid-range SUM drop (omits an interior month, not the last)
      * one line-item row with an off-by-small multiplier
      * a subtotal that adds a wrong adjacent row
      * a hardcoded near-correct total
      * an absolute/relative ref slip that only bites in one cell
"""
from __future__ import annotations

import json
import os
import random

import openpyxl
from openpyxl.utils import get_column_letter

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
FIRST_COL = 2
LAST_COL = FIRST_COL + 11         # M
TOTAL_COL = LAST_COL + 1          # N
N_ITEMS = 15
ROW_START = 3
ROW_TOTAL = ROW_START + N_ITEMS   # grand total row


def col(i: int) -> str:
    return get_column_letter(i)


def main() -> None:
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OpEx"
    ws["A1"] = "Operating Expense Budget (15 line items x 12 months)"
    t = col(TOTAL_COL)
    ws[f"{t}1"] = "FY Total"
    for idx, c in enumerate(range(FIRST_COL, LAST_COL + 1)):
        ws[f"{col(c)}1"] = MONTHS[idx]

    # base monthly values per line item (row 2 is a hidden driver? no -- direct)
    bases = [random.randint(800, 5000) for _ in range(N_ITEMS)]
    growth = 1.03

    truth = []

    for i in range(N_ITEMS):
        row = ROW_START + i
        ws[f"A{row}"] = f"Line Item {i+1}"
        base = bases[i]
        for j, c in enumerate(range(FIRST_COL, LAST_COL + 1)):
            letter = col(c)
            if c == FIRST_COL:
                ws[f"{letter}{row}"] = base
            else:
                prev = col(c - 1)
                # each month = prev * growth, as a formula
                ws[f"{letter}{row}"] = f"={prev}{row}*{growth}"
        # FY total per line item
        ws[f"{t}{row}"] = f"=SUM({col(FIRST_COL)}{row}:{col(LAST_COL)}{row})"

    # grand total row = SUM of line items per column
    ws[f"A{ROW_TOTAL}"] = "TOTAL OpEx"
    for c in range(FIRST_COL, TOTAL_COL + 1):
        letter = col(c)
        ws[f"{letter}{ROW_TOTAL}"] = (
            f"=SUM({letter}{ROW_START}:{letter}{ROW_TOTAL-1})"
        )

    # ---- inject SUBTLE errors ----

    # HARD-1: mid-range SUM drop on Line Item 6's FY total (omits interior month Jun=G)
    r = ROW_START + 5
    cell = f"{t}{r}"
    ws[cell] = f"=SUM({col(FIRST_COL)}{r}:{col(LAST_COL)}{r})-{col(7)}{r}"  # subtract Jun
    truth.append({"cell": cell, "class": "E1",
                  "desc": "Line 6 FY total silently omits June (interior month).",
                  "correct_formula": f"=SUM({col(FIRST_COL)}{r}:{col(LAST_COL)}{r})"})

    # HARD-2: Line Item 10 uses growth 1.30 in ONE month (Sep=J), plausible-looking
    r = ROW_START + 9
    jcol = col(10)  # J = Sep
    prev = col(9)
    cell = f"{jcol}{r}"
    ws[cell] = f"={prev}{r}*1.30"  # neighbors use *1.03
    truth.append({"cell": cell, "class": "E3",
                  "desc": "Line 10 Sep uses *1.30 vs row's *1.03.",
                  "correct_formula": f"={prev}{r}*1.03"})

    # HARD-3: grand total for Aug (I) sums one row too few
    icol = col(9)  # I = Aug
    cell = f"{icol}{ROW_TOTAL}"
    ws[cell] = f"=SUM({icol}{ROW_START}:{icol}{ROW_TOTAL-2})"  # stops one row short
    truth.append({"cell": cell, "class": "E1",
                  "desc": "Aug grand total omits the last line item (row short).",
                  "correct_formula": f"=SUM({icol}{ROW_START}:{icol}{ROW_TOTAL-1})"})

    # HARD-4: FY grand total (N total row) hardcoded to a near-correct literal
    cell = f"{t}{ROW_TOTAL}"
    ws[cell] = 615000  # plausible-looking round number
    truth.append({"cell": cell, "class": "E7",
                  "desc": "FY grand total hardcoded (near-correct literal).",
                  "correct_formula": f"=SUM({t}{ROW_START}:{t}{ROW_TOTAL-1})"})

    os.makedirs("data/sheets", exist_ok=True)
    os.makedirs("data/truth", exist_ok=True)
    wb.save("data/sheets/model_02_hard.xlsx")
    with open("data/truth/model_02_hard.json", "w") as f:
        json.dump({"sheet": "OpEx", "errors": truth}, f, indent=2)
    print(f"wrote data/sheets/model_02_hard.xlsx with {len(truth)} subtle errors")
    for e in truth:
        print(f"  {e['cell']} [{e['class']}] {e['desc']}")


if __name__ == "__main__":
    main()
